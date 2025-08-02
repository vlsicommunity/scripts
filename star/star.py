#------------------------------------------
# Front End of STAR - STA Reporting
# Author: Harish Kumar R - nxf67811
#------------------------------------------

import sys
starCmd = sys.argv[1]

def get_real_name(var_userid):
  var_child = subprocess.Popen( \
                ['getent', 'passwd', var_userid], \
                  stdout=subprocess.PIPE)
  var_name  = var_child.communicate()[0].decode('utf8').split(":")[4]
  return var_name

def sendEmail(toIds,
              ccIds,
              bccIds,
              sub,
              mailBody):
 
  var_child = subprocess.Popen( \
                ['/pkg/OSS-mutt-/2.2.10/x86_64-linux4.18-glibc2.28/bin/mutt', \
                 '-e set content_type=text/html', \
                 "-c " + ccIds, \
                 "-b " + bccIds, \
                 " " + toIds, \
                 "-s " + sub], \
                 stdin=subprocess.PIPE, \
                 stdout=subprocess.PIPE)
  var_child.communicate(mail_body.encode('utf8'))

#------------------------------------------
# Set STAR config variables
#------------------------------------------
import yaml
from yaml.loader import SafeLoader

if starCmd == "rc":

  import os
  starResourceDir = os.path.dirname(os.path.realpath(__file__))
  cwd = sys.argv[2]

  with open(starResourceDir + "/star_config.yaml") as f:
    starConfigInfo = yaml.load(f, Loader=SafeLoader)
  if starConfigInfo["misc_settings"]["output_dir"] == "":
    from datetime import datetime
    starOutDir = cwd + '/star_' + datetime.today().strftime('y%Ym%md%d')
  else:
    starOutDir = starConfigInfo["misc_settings"]["output_dir"]

  fp = open(cwd + "/.sc","w")
  fp.write('set ::star::config' + ' [dict create]' + "\n")
  fp.write('set ::star::unixOutDir ' + starOutDir + "\n")

  for cat in starConfigInfo.keys():
    for key in starConfigInfo[cat].keys():
      if isinstance(starConfigInfo[cat][key], list):
        fp.write('dict set ::star::config ' + 
              cat + " " + key + " ")
        fp.write('[list ')
        for item in starConfigInfo[cat][key]:
          fp.write(item + " ")
        fp.write(']' + "\n")
      elif isinstance(starConfigInfo[cat][key], str):
        if starConfigInfo[cat][key] == "":
          fp.write('dict set ::star::config ' + cat + 
                                " " + key + " " + '""' 
                                + "\n")
        else:
          fp.write('dict set ::star::config ' + cat + 
            " " + key + " " + starConfigInfo[cat][key] 
            + "\n")
      else:
        fp.write('dict set ::star::config ' + cat + 
            " " + key + " " + str(starConfigInfo[cat][key]) 
            + "\n")
  fp.close()

if starCmd == "pub": 

  import os
  starResourceDir = os.path.dirname(os.path.realpath(__file__))
  with open(starResourceDir + "/star_config.yaml") as f:
    starConfigInfo = yaml.load(f, Loader=SafeLoader)

  # Read STAR Timing Info
  starEngViewDir = sys.argv[2]
  with open(starEngViewDir + "/.star_tcl2py_miscInfo.yaml") as f:
    temp = yaml.load(f, Loader=SafeLoader)
  miscInfo = temp["miscInfo"]  

  with open(starEngViewDir + "/.star_tcl2py_pvtInfo.yaml") as f:
    pvtInfo = yaml.load(f, Loader=SafeLoader)

  with open(starEngViewDir + "/.star_tcl2py_timingInfo.yaml") as f:
    starTimingInfo = yaml.load(f, Loader=SafeLoader)
  timingInfo = starTimingInfo["timingInfo"]

  #print(miscInfo)
  buildTag = miscInfo["build_tag"]
  designTop = miscInfo["design_name"]
  sub = "STAR report for design: " + designTop + "; STA run tag: " + buildTag

  # Send email
  import subprocess

  mail_body = "<html>"
  mail_body += "<head><STAR></title>"
  mail_body += "</head>"
  mail_body += '<body class="clean-body u_body" style="margin: 0;padding: 0;-webkit-text-size-adjust: 100%;background-color: #ffffff;color: #000000">'
  #mail_body += '<table>'
  #mail_body += '<tr>'
  #for modeCat in timingInfo:
  #  mail_body += "<p> ModeCat : " + modeCat + "</p>"
  #  for delay in timingInfo[modeCat]:
  #    mail_body += "<p> Delay : " + delay + "</p>"
  #    for mainCat in timingInfo[modeCat][delay]:
  #      mail_body += "<p> Cat : " + mainCat + "</p>"
  #      for mode in timingInfo[modeCat][delay][mainCat]["mode_stats"]:
  #        mail_body += "<p> mode : " + mode + "</p>"
  #        mail_body += "<p> wns : " + str(timingInfo[modeCat][delay][mainCat]["mode_stats"][mode]["wns"]) + "</p>"
  #mail_body += '<table>'
  #mail_body += """
  #<section style="font-size: 14px; line-height: 140%; background-color: lavender;">
  #  <p style="font-size: 14px; line-height: 140%; text-align: center;">
  #    <strong>
  #    <span style="font-size: 32px; color: #000000;"> STAR </span>
  #    </strong>
  #  </p>
  #  <p style="font-size: 14px; line-height: 140%; text-align: center;">
  #    <strong>
  #    <span style="font-size: 18px; color: #000000;">STA-Reporting System</span>
  #    </strong>
  #  </p>
  #</section>
  #"""
  mail_body += """<div style="background:#EBF5FB; margin-top:10px; width:100%; height:60px; display:block !important;">"""
  mail_body += """  <span style="font-weight:300; font-size: 40px; line-height: 125%; font-family: Verdana; text-align:center;">&#9734 </span>"""
  mail_body += """  <span style="font-weight:300; font-size: 20px; line-height: 125%; font-family: Verdana; text-align:center;">STA-Reporting </span>"""
  mail_body += """</div>"""
  mail_body += "<p>Reports dir : " + starEngViewDir + " </p>"
  mail_body += "</body>"
  mail_body += "</html>"

  toIds = ",".join(starConfigInfo["mail_settings"]["to_list"])
  if starConfigInfo["mail_settings"]["cc_list"] is None or starConfigInfo["mail_settings"]["bcc_list"] == "":
    ccIds = ""
  else: 
    ccIds = ",".join(starConfigInfo["mail_settings"]["cc_list"])
  if starConfigInfo["mail_settings"]["bcc_list"] is None or starConfigInfo["mail_settings"]["bcc_list"] == "":
    bccIds = "harishkumar.rajendrababu@nxp.com"
  else:
    starConfigInfo["mail_settings"]["bcc_list"].append("harishkumar.rajendrababu@nxp.com")
    bccIds = ",".join(starConfigInfo["mail_settings"]["bcc_list"])
  #print(bccIds)

  import openpyxl
  wb = openpyxl.Workbook()
  for modeGrp in timingInfo:
    for delay in timingInfo[modeGrp]:
      sheetName = modeGrp + "_" + delay
      wb.create_sheet(sheetName)
      sheet = wb[sheetName]
      rowIndex = 0
      colIndex = 0
      #print header for overview table
      tableBannerlen = 1
      c1 = sheet.cell(row = 1, column = 1)
      c1.value = "ANKIT"
  wb.save(starEngViewDir + "/eview.xlsx")
  attachment = starEngViewDir + "/eview.xlsx"
  sendEmail(toIds,ccIds,bccIds,sub,mail_body)
